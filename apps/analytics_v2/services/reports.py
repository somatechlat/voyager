"""Report generation service — PDF, CSV, Excel, and JSON output.

Handles report rendering pipeline: data collection from ClickHouse,
visualization generation, format-specific output, and delivery.
"""

from __future__ import annotations

import csv
import io
import json
import logging
from datetime import datetime
from typing import Any

from apps.analytics_v2.services.dashboards import compute_date_range

logger = logging.getLogger(__name__)


def generate_report(
    template_config: dict[str, Any],
    output_format: str,
    date_range: dict[str, str],
    filters: dict[str, Any],
    tenant_id: str,
) -> dict[str, Any]:
    """Generate a report from template configuration.

    Args:
        template_config: Report template config with metrics, dimensions, filters.
        output_format: Output format (pdf, csv, excel, json).
        date_range: Date range dict with start/end or preset.
        filters: Additional filters to apply.
        tenant_id: Tenant scope.

    Returns:
        Dict with job_id, status, file_path, row_count, etc.
    """
    start_dt, end_dt = compute_date_range(
        date_range.get("preset"),
        date_range.get("start"),
        date_range.get("end"),
    )

    logger.info(
        "Generating %s report for tenant %s, range %s to %s",
        output_format,
        tenant_id,
        start_dt,
        end_dt,
    )

    # 1. Collect data from ClickHouse
    data = _collect_report_data(template_config, start_dt, end_dt, filters, tenant_id)

    # 2. Format output
    if output_format == "csv":
        result = _format_csv(data, template_config)
    elif output_format == "json":
        result = _format_json(data, template_config, start_dt, end_dt, filters)
    elif output_format == "excel":
        result = _format_excel(data, template_config)
    elif output_format == "pdf":
        result = _format_pdf(data, template_config)
    else:
        result = {"status": "failed", "error": f"Unsupported format: {output_format}"}

    result["format"] = output_format
    result["row_count"] = len(data.get("rows", []))
    result["generated_at"] = datetime.utcnow().isoformat()

    return result


def _collect_report_data(
    config: dict[str, Any],
    start_dt: datetime,
    end_dt: datetime,
    filters: dict[str, Any],
    tenant_id: str,
) -> dict[str, Any]:
    """Collect report data by executing queries against ClickHouse.

    Args:
        config: Report configuration.
        start_dt: Start datetime.
        end_dt: End datetime.
        filters: Query filters.
        tenant_id: Tenant scope.

    Returns:
        Dict with columns and rows.
    """
    metrics = config.get("metrics", [])
    dimensions = config.get("dimensions", [])
    columns = dimensions + [m["alias"] if isinstance(m, dict) else m for m in metrics]

    rows = []
    try:
        from django.db import connections

        ch = connections.get("clickhouse")
        where = f"tenant_id = '{tenant_id}' AND event_date BETWEEN '{start_dt.date()}' AND '{end_dt.date()}'"
        platform_filter = filters.get("platform", "")
        if platform_filter:
            where += f" AND platform = '{platform_filter}'"

        dim_select = ", ".join(dimensions) if dimensions else "platform"
        metric_selects = []
        for m in metrics:
            if isinstance(m, dict):
                agg = m.get("aggregation", "sum")
                name = m.get("metric", "")
                alias = m.get("alias", name)
                metric_selects.append(f"{agg}(metric_value) as {alias}")
            else:
                metric_selects.append(f"sum(metric_value) as {m}")
        metric_select = (
            ", ".join(metric_selects) if metric_selects else "sum(metric_value) as total"
        )

        sql = f"""
            SELECT {dim_select}, {metric_select}
            FROM analytics_events
            WHERE {where}
            GROUP BY {dim_select}
            ORDER BY total DESC
            LIMIT 10000
        """
        with ch.cursor() as cursor:
            cursor.execute(sql)
            cols = [desc[0] for desc in cursor.description]
            rows = [dict(zip(cols, row)) for row in cursor.fetchall()]
    except Exception as exc:
        logger.warning("ClickHouse query failed: %s", exc)
        rows = []

    return {"columns": columns, "rows": rows, "metrics": metrics, "dimensions": dimensions}


def _format_csv(
    data: dict[str, Any],
    config: dict[str, Any],
) -> dict[str, Any]:
    """Format report data as CSV with UTF-8 BOM for Excel compatibility.

    Args:
        data: Collected report data with columns and rows.
        config: Report configuration.

    Returns:
        Dict with content (bytes) and content_type.
    """
    output = io.StringIO()
    writer = csv.writer(output)
    # UTF-8 BOM
    bom = "\ufeff"
    rows = data.get("rows", [])
    if rows:
        headers = list(rows[0].keys())
        writer.writerow([bom + headers[0]] + headers[1:])
        for row in rows:
            writer.writerow([row.get(h, "") for h in headers])
    else:
        writer.writerow(["No data"])

    return {
        "status": "completed",
        "content": output.getvalue().encode("utf-8"),
        "content_type": "text/csv; charset=utf-8",
        "filename": f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv",
    }


def _format_json(
    data: dict[str, Any],
    config: dict[str, Any],
    start_dt: datetime,
    end_dt: datetime,
    filters: dict[str, Any],
) -> dict[str, Any]:
    """Format report data as structured JSON with metadata.

    Args:
        data: Collected report data.
        config: Report configuration.
        start_dt: Report start date.
        end_dt: Report end date.
        filters: Applied filters.

    Returns:
        Dict with content (bytes) and content_type.
    """
    payload = {
        "metadata": {
            "generated_at": datetime.utcnow().isoformat(),
            "date_range": {"start": start_dt.isoformat(), "end": end_dt.isoformat()},
            "filters": filters,
            "metrics": config.get("metrics", []),
            "dimensions": config.get("dimensions", []),
            "total_rows": len(data.get("rows", [])),
        },
        "data": data.get("rows", []),
    }
    content = json.dumps(payload, indent=2, default=str).encode("utf-8")
    return {
        "status": "completed",
        "content": content,
        "content_type": "application/json",
        "filename": f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.json",
    }


def _format_excel(
    data: dict[str, Any],
    config: dict[str, Any],
) -> dict[str, Any]:
    """Format report data as Excel (.xlsx) with multiple sheets.

    Args:
        data: Collected report data.
        config: Report configuration.

    Returns:
        Dict with content (bytes) and content_type.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        logger.warning("openpyxl not available; falling back to CSV within Excel container")
        return _format_csv(data, config)

    wb = openpyxl.Workbook()

    # Data sheet
    ws = wb.active
    ws.title = "Report Data"
    rows = data.get("rows", [])
    if rows:
        headers = list(rows[0].keys())
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, row in enumerate(rows, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
    else:
        ws.cell(row=1, column=1, value="No data available")

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Metric")
    ws2.cell(row=1, column=2, value="Total")
    if rows:
        numeric_cols = [k for k, v in rows[0].items() if isinstance(v, (int, float))]
        for idx, col in enumerate(numeric_cols, 2):
            total = sum(row.get(col, 0) for row in rows)
            ws2.cell(row=idx, column=1, value=col)
            ws2.cell(row=idx, column=2, value=total)

    output = io.BytesIO()
    wb.save(output)
    return {
        "status": "completed",
        "content": output.getvalue(),
        "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "filename": f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
    }


def _format_pdf(
    data: dict[str, Any],
    config: dict[str, Any],
) -> dict[str, Any]:
    """Format report as PDF with header, data tables, and footer.

    Args:
        data: Collected report data.
        config: Report configuration.

    Returns:
        Dict with content (bytes) and content_type.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        logger.warning("reportlab not available; falling back to CSV")
        return _format_csv(data, config)

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    elements = []

    # Title
    from reportlab.lib.styles import getSampleStyleSheet

    styles = getSampleStyleSheet()
    elements.append(Paragraph(config.get("title", "Analytics Report"), styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(
        Paragraph(
            f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 20))

    # Data table
    rows = data.get("rows", [])
    if rows:
        headers = list(rows[0].keys())
        table_data = [headers] + [[str(row.get(h, "")) for h in headers] for row in rows[:500]]
        table = Table(table_data)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("FONTSIZE", (0, 1), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [colors.white, colors.HexColor("#f0f0f0")],
                    ),
                ]
            )
        )
        elements.append(table)
    else:
        elements.append(Paragraph("No data available.", styles["Normal"]))

    doc.build(elements)
    return {
        "status": "completed",
        "content": output.getvalue(),
        "content_type": "application/pdf",
        "filename": f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf",
    }


def deliver_report(
    report_result: dict[str, Any],
    delivery_config: dict[str, Any],
) -> dict[str, Any]:
    """Deliver a generated report via configured channels.

    Args:
        report_result: Output from generate_report().
        delivery_config: Dict with delivery methods and parameters.

    Returns:
        Dict with delivery status per channel.
    """
    statuses = {}
    for method, cfg in delivery_config.items():
        try:
            if method == "email":
                _deliver_email(report_result, cfg)
                statuses["email"] = "sent"
            elif method == "slack":
                _deliver_slack(report_result, cfg)
                statuses["slack"] = "sent"
            elif method == "webhook":
                _deliver_webhook(report_result, cfg)
                statuses["webhook"] = "sent"
            elif method == "s3":
                _deliver_s3(report_result, cfg)
                statuses["s3"] = "uploaded"
            else:
                statuses[method] = "skipped"
        except Exception as exc:
            logger.error("Delivery failed for %s: %s", method, exc)
            statuses[method] = f"failed: {exc}"

    return statuses


def _deliver_email(result: dict[str, Any], cfg: dict[str, Any]) -> None:
    """Send report via email."""
    from django.core.mail import EmailMessage

    msg = EmailMessage(
        subject=cfg.get("subject", "Analytics Report"),
        body=cfg.get("body", "Please find your report attached."),
        from_email=cfg.get("from"),
        to=cfg.get("recipients", []),
    )
    content = result.get("content", b"")
    filename = result.get("filename", "report.csv")
    msg.attach(filename, content, result.get("content_type", "text/csv"))
    msg.send()


def _deliver_slack(result: dict[str, Any], cfg: dict[str, Any]) -> None:
    """Send report summary via Slack webhook."""

    import httpx

    webhook_url = cfg.get("webhook_url", "")
    if not webhook_url:
        return
    payload = {
        "text": f"Report generated: {result.get('filename')} ({result.get('row_count', 0)} rows)",
        "channel": cfg.get("channel", "#reports"),
    }
    httpx.post(webhook_url, json=payload, timeout=30)


def _deliver_webhook(result: dict[str, Any], cfg: dict[str, Any]) -> None:
    """Send report via custom webhook."""
    import httpx

    url = cfg.get("url", "")
    if not url:
        return
    httpx.post(
        url,
        json={
            "filename": result.get("filename"),
            "row_count": result.get("row_count"),
            "format": result.get("format"),
            "generated_at": result.get("generated_at"),
        },
        timeout=30,
    )


def _deliver_s3(result: dict[str, Any], cfg: dict[str, Any]) -> None:
    """Upload report to S3."""
    import boto3

    s3 = boto3.client("s3")
    s3.put_object(
        Bucket=cfg.get("bucket", "voyager-reports"),
        Key=result.get("filename", "report.csv"),
        Body=result.get("content", b""),
        ContentType=result.get("content_type", "text/csv"),
    )
